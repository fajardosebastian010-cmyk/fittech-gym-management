from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.db import models
from django.db.models import Count, Sum
from django.db.models import Q
from datetime import timedelta, datetime
from .models import Usuario, Membresia, Cliente, Asistencia, HistorialMembresia, Pago, Bono
from .dao import UsuarioDAO, MembresiaDAO, ClienteDAO, AsistenciaDAO, PagoDAO
from .email_utils import EmailService
import openpyxl
from django.http import HttpResponse, JsonResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import pandas as pd
from datetime import date
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def actualizar_estados_clientes():
    """Actualiza automáticamente el estado de los clientes según su fecha de vencimiento"""
    from django.utils import timezone
    hoy = timezone.now().date()
    
    # Clientes con membresía vencida pero estado activo
    clientes_vencidos = Cliente.objects.filter(
        estado='activo',
        fecha_fin_membresia__lt=hoy
    )
    clientes_vencidos.update(estado='inactivo')
    
    # Clientes con membresía vigente pero estado inactivo (renovaciones)
    clientes_activos = Cliente.objects.filter(
        estado='inactivo',
        fecha_fin_membresia__gte=hoy
    )
    clientes_activos.update(estado='activo')
# ============= DECORADORES PERSONALIZADOS =============
def es_administrador(user):
    return user.is_authenticated and user.rol == 'administrador'

def es_empleado_o_admin(user):
    return user.is_authenticated and user.rol in ['empleado', 'administrador']

# ============= LOGIN Y AUTENTICACIÓN =============
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        correo = request.POST.get('correo')
        password = request.POST.get('password')
        user = authenticate(request, username=correo, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Bienvenido {user.nombre}')
            return redirect('dashboard')
        else:
            messages.error(request, 'Credenciales incorrectas')
    
    return render(request, 'login.html')

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, 'Sesión cerrada correctamente')
    return redirect('login')

# ============= DASHBOARD CON GRÁFICAS - CORREGIDO =============
@login_required
def dashboard(request):
    actualizar_estados_clientes()
    # Estadísticas generales
    total_membresias = Membresia.objects.filter(activa=True).count()
    total_clientes = Cliente.objects.count()
    total_usuarios = Usuario.objects.count()
    asistencias_hoy = AsistenciaDAO.contar_asistencias_dia()

    hoy = timezone.now().date()

    # Clientes por vencer en los próximos 7 días
    
    fecha_limite = hoy + timedelta(days=7)
    clientes_por_vencer = Cliente.objects.filter(
    fecha_fin_membresia__gte=hoy,
    fecha_fin_membresia__lte=fecha_limite,
    estado='activo'
    )

    # Clientes con membresías vencidas
    clientes_vencidos = Cliente.objects.filter(
    estado='inactivo'
    )

    
    pagos_pendientes = Pago.objects.filter(estado='pendiente').count()
    
    # CORRECCIÓN DEFINITIVA - Usar rangos de datetime
    ahora = timezone.now()
    inicio_dia = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
    fin_dia = ahora.replace(hour=23, minute=59, second=59, microsecond=999999)
    inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    hoy_date = ahora.date()
    
    # Ingresos de hoy usando rangos de datetime
    ingresos_hoy = Pago.objects.filter(
        fecha_pago__gte=inicio_dia,
        fecha_pago__lte=fin_dia,
        estado='validado'
    ).aggregate(Sum('monto'))['monto__sum'] or 0
    
    # Ingresos del mes usando rangos de datetime
    ingresos_mes_actual = Pago.objects.filter(
        fecha_pago__gte=inicio_mes,
        fecha_pago__lte=ahora,
        estado='validado'
    ).aggregate(Sum('monto'))['monto__sum'] or 0
    
    # Clientes activos vs inactivos
    clientes_activos = Cliente.objects.filter(estado='activo').count()
    clientes_inactivos = Cliente.objects.filter(estado='inactivo').count()
    
    # Ingresos de los últimos 6 meses
    meses_labels = []
    meses_ingresos = []
    
    for i in range(5, -1, -1):
        fecha = hoy_date - timedelta(days=30*i)
        inicio_mes_iter = fecha.replace(day=1)
        
        if fecha.month == 12:
            fin_mes_iter = fecha.replace(day=31)
        else:
            siguiente_mes = fecha.replace(day=28) + timedelta(days=4)
            fin_mes_iter = siguiente_mes - timedelta(days=siguiente_mes.day)
        
        # Convertir a datetime para la consulta
        inicio_dt = timezone.make_aware(datetime.combine(inicio_mes_iter, datetime.min.time()))
        fin_dt = timezone.make_aware(datetime.combine(fin_mes_iter, datetime.max.time()))
        
        ingresos_mes = Pago.objects.filter(
            fecha_pago__gte=inicio_dt,
            fecha_pago__lte=fin_dt,
            estado='validado'
        ).aggregate(Sum('monto'))['monto__sum'] or 0
        
        meses_es = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        mes_nombre = meses_es[fecha.month]
        
        meses_labels.append(f"{mes_nombre} {fecha.year}")
        meses_ingresos.append(float(ingresos_mes))
    
    # Asistencias de los últimos 7 días
    dias_labels = []
    dias_asistencias = []
    
    for i in range(6, -1, -1):
        fecha = hoy_date - timedelta(days=i)
        asistencias_dia = Asistencia.objects.filter(fecha=fecha).count()
        dias_labels.append(fecha.strftime('%d/%m'))
        dias_asistencias.append(asistencias_dia)
    
    # Distribución de clientes por membresía
    membresias_distribucion = Cliente.objects.filter(estado='activo').values('membresia_actual__nombre').annotate(total=Count('documento')).order_by('-total')
    membresias_labels = [item['membresia_actual__nombre'] or 'Sin membresía' for item in membresias_distribucion]
    membresias_valores = [item['total'] for item in membresias_distribucion]
    
    # Pagos por método de pago (últimos 30 días)
    fecha_hace_30 = ahora - timedelta(days=30)
    pagos_por_metodo = Pago.objects.filter(
        fecha_pago__gte=fecha_hace_30,
        fecha_pago__lte=ahora,
        estado='validado'
    ).values('metodo_pago').annotate(total=Count('id'), monto_total=Sum('monto')).order_by('-monto_total')
    
    metodos_labels = [dict(Pago.METODOS_PAGO).get(item['metodo_pago'], item['metodo_pago']) for item in pagos_por_metodo]
    metodos_valores = [float(item['monto_total']) for item in pagos_por_metodo]
    
    context = {
        'total_membresias': total_membresias,
        'total_clientes': total_clientes,
        'total_usuarios': total_usuarios,
        'asistencias_hoy': asistencias_hoy,
        'clientes_por_vencer': clientes_por_vencer,
        'clientes_vencidos': clientes_vencidos,
        'pagos_pendientes': pagos_pendientes,
        'ingresos_hoy': float(ingresos_hoy),
        'clientes_activos': clientes_activos,
        'clientes_inactivos': clientes_inactivos,
        'ingresos_mes_actual': float(ingresos_mes_actual),
        'usuario': request.user,
        
        # Datos para gráficas
        'meses_labels': meses_labels,
        'meses_ingresos': meses_ingresos,
        'dias_labels': dias_labels,
        'dias_asistencias': dias_asistencias,
        'membresias_labels': membresias_labels,
        'membresias_valores': membresias_valores,
        'metodos_labels': metodos_labels,
        'metodos_valores': metodos_valores,
    }
    return render(request, 'dashboard.html', context)

# ============= MEMBRESÍAS =============
@login_required
@user_passes_test(es_administrador)
def membresias_listar(request):
    membresias = MembresiaDAO.obtener_todas()
    return render(request, 'membresias/listar.html', {'membresias': membresias})

@login_required
@user_passes_test(es_administrador)
def membresias_crear(request):
    if request.method == 'POST':
        try:
            datos = {
                'nombre': request.POST.get('nombre'),
                'duracion_dias': int(request.POST.get('duracion_dias')),
                'precio': float(request.POST.get('precio')),
                'descripcion': request.POST.get('descripcion', ''),
            }
            MembresiaDAO.crear(datos)
            messages.success(request, 'Membresía creada exitosamente')
            return redirect('membresias_listar')
        except Exception as e:
            messages.error(request, f'Error al crear la membresía: {str(e)}')
    
    return render(request, 'membresias/crear.html')

@login_required
@user_passes_test(es_administrador)
def membresias_editar(request, id):
    membresia = get_object_or_404(Membresia, id=id)
    
    if request.method == 'POST':
        try:
            datos = {
                'nombre': request.POST.get('nombre'),
                'duracion_dias': int(request.POST.get('duracion_dias')),
                'precio': float(request.POST.get('precio')),
                'descripcion': request.POST.get('descripcion', ''),
            }
            MembresiaDAO.actualizar(id, datos)
            messages.success(request, 'Membresía actualizada exitosamente')
            return redirect('membresias_listar')
        except Exception as e:
            messages.error(request, f'Error al actualizar la membresía: {str(e)}')
    
    return render(request, 'membresias/editar.html', {'membresia': membresia})

@login_required
@user_passes_test(es_administrador)
def membresias_eliminar(request, id):
    try:
        MembresiaDAO.eliminar(id)
        messages.success(request, 'Membresía eliminada exitosamente')
    except Exception as e:
        messages.error(request, f'Error al eliminar la membresía: {str(e)}')
    return redirect('membresias_listar')

@login_required
@user_passes_test(es_administrador)
def membresias_ver(request, id):
    membresia = get_object_or_404(Membresia, id=id)
    return render(request, 'membresias/ver.html', {'membresia': membresia})

# ============= CLIENTES CON FILTRO DE BÚSQUEDA Y BONOS =============
@login_required
def clientes_listar(request):
    busqueda = request.GET.get('busqueda', '')
    estado_filtro = request.GET.get('estado', 'todos')
    
    # Base queryset
    clientes = Cliente.objects.all()
    
    # Filtro de búsqueda por texto
    if busqueda:
        clientes = clientes.filter(
            Q(documento__icontains=busqueda) |
            Q(nombres__icontains=busqueda) |
            Q(apellidos__icontains=busqueda) |
            Q(email__icontains=busqueda) |
            Q(celular__icontains=busqueda)
        )
    
    # Filtro por estado
    if estado_filtro == 'activo':
        clientes = clientes.filter(estado='activo')
    elif estado_filtro == 'inactivo':
        clientes = clientes.filter(estado='inactivo')
    elif estado_filtro == 'pendiente':
        clientes = clientes.filter(estado='pendiente')
    elif estado_filtro == 'por_vencer':
        # Clientes activos cuya membresía vence en los próximos 7 días
        hoy = datetime.now().date()
        fecha_limite = hoy + timedelta(days=7)
        clientes = clientes.filter(
            estado='activo',
            fecha_fin_membresia__gte=hoy,
            fecha_fin_membresia__lte=fecha_limite
        )
    
    # Ordenar por fecha de vencimiento
    clientes = clientes.order_by('fecha_fin_membresia')
    
    return render(request, 'clientes/listar.html', {
        'clientes': clientes,
        'busqueda': busqueda,
        'estado_filtro': estado_filtro
    })
    
    return render(request, 'clientes/listar.html', context)

@login_required
def clientes_crear(request):
    if request.method == 'POST':
        try:
            # VALIDACIÓN DE EDAD MÍNIMA (solo si se proporciona fecha de nacimiento)
            fecha_nacimiento_str = request.POST.get('fecha_nacimiento')
            fecha_nacimiento = None
            
            if fecha_nacimiento_str:
                fecha_nacimiento = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
                hoy = timezone.now().date()
                edad = hoy.year - fecha_nacimiento.year - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
                
                if edad < 14:
                    messages.error(request, f'❌ El cliente debe tener al menos 14 años para registrarse. Edad actual: {edad} años')
                    membresias = MembresiaDAO.obtener_todas()
                    return render(request, 'clientes/crear.html', {'membresias': membresias})
            
            # Obtener membresía
            membresia = get_object_or_404(Membresia, id=request.POST.get('membresia'))
            fecha_inicio = timezone.now().date()
            fecha_fin = fecha_inicio + timedelta(days=membresia.duracion_dias)
            
            # BONO DE BIENVENIDA (máx 3 días)
            bono_dias = request.POST.get('bono_dias', '0')
            dias_extra = 0
            if bono_dias != '0':
                dias_extra = int(bono_dias)
                if dias_extra > 3:
                    dias_extra = 3
                fecha_fin += timedelta(days=dias_extra)
            
            # Estado inicial: pendiente (hasta que se valide el pago)
            estado_cliente = 'pendiente'
            
            # REGISTRO DEL CLIENTE CON TIPO DE DOCUMENTO
            peso_str = request.POST.get('peso')
            peso = float(peso_str) if peso_str else None
            
            datos = {
                'tipo_documento': request.POST.get('tipo_documento', 'CC'),  # ← NUEVO CAMPO
                'documento': request.POST.get('documento'),
                'nombres': request.POST.get('nombres'),
                'apellidos': request.POST.get('apellidos'),
                'peso': peso,  # Puede ser None
                'fecha_nacimiento': fecha_nacimiento,  # Puede ser None
                'email': request.POST.get('email') or None,
                'celular': request.POST.get('celular') or None,
                'membresia_actual': membresia,
                'fecha_inicio_membresia': fecha_inicio,
                'fecha_fin_membresia': fecha_fin,
                'estado': estado_cliente,
            }
            cliente = ClienteDAO.crear(datos)
            
            # REGISTRAR BONO SI SE OTORGÓ
            if dias_extra > 0:
                tipo_bono_map = {1: '1_dia', 2: '2_dias', 3: '3_dias'}
                Bono.objects.create(
                    cliente=cliente,
                    tipo_bono=tipo_bono_map.get(dias_extra, '1_dia'),
                    dias_regalo=dias_extra,
                    motivo='Regalo de bienvenida',
                    usuario_otorgo=request.user,
                    aplicado=True,
                    fecha_aplicado=timezone.now()
                )
            
            # REGISTRO DEL PAGO si se llenó método y monto
            metodo_pago = request.POST.get('metodo_pago')
            monto = request.POST.get('monto')
            comprobante = request.POST.get('comprobante', '')
            observaciones = request.POST.get('observaciones', '')
            
            if metodo_pago and monto:
                datos_pago = {
                    'cliente': cliente,
                    'membresia': membresia,
                    'concepto': f"Pago inicial - Membresía {membresia.nombre}",
                    'tipo_pago': 'membresia',
                    'monto': float(monto),
                    'metodo_pago': metodo_pago,
                    'comprobante': comprobante,
                    'observaciones': observaciones,
                    'usuario_registro': request.user,
                    'estado': 'pendiente',
                }
                PagoDAO.crear(datos_pago)
                # Cliente queda en estado 'pendiente' (esperando validación del pago)
                cliente.estado = 'pendiente'
                cliente.save()
            
            # REGISTRO EN HISTORIAL
            HistorialMembresia.objects.create(
                cliente=cliente,
                membresia=membresia,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                precio_pagado=membresia.precio
            )
            
            # EMAIL BIENVENIDA
            try:
                EmailService.enviar_email_bienvenida(cliente)
                mensaje_exito = f'✅ Cliente registrado exitosamente'
                if metodo_pago and monto:
                    mensaje_exito += ' con pago pendiente de validación'
                if dias_extra > 0:
                    mensaje_exito += f' (incluye {dias_extra} día{"s" if dias_extra > 1 else ""} de regalo)'
                mensaje_exito += '. Email de bienvenida enviado.'
                messages.success(request, mensaje_exito)
            except Exception as email_error:
                mensaje_exito = f'✅ Cliente registrado exitosamente'
                if metodo_pago and monto:
                    mensaje_exito += ' con pago pendiente de validación'
                if dias_extra > 0:
                    mensaje_exito += f' (incluye {dias_extra} día{"s" if dias_extra > 1 else ""} de regalo)'
                mensaje_exito += '. No se pudo enviar el email de bienvenida.'
                messages.warning(request, mensaje_exito)

            return redirect('clientes_listar')
            
        except Exception as e:
            messages.error(request, f'❌ Error al registrar el cliente: {str(e)}')
            import traceback
            print(traceback.format_exc())  # Para debugging
    
    membresias = MembresiaDAO.obtener_todas()
    return render(request, 'clientes/crear.html', {'membresias': membresias})

@login_required
def clientes_editar(request, documento):
    cliente = get_object_or_404(Cliente, documento=documento)
    
    if request.method == 'POST':
        try:
            datos = {
                'nombres': request.POST.get('nombres'),
                'apellidos': request.POST.get('apellidos'),
                'peso': float(request.POST.get('peso')),
                'fecha_nacimiento': request.POST.get('fecha_nacimiento'),
                'email': request.POST.get('email'),
                'celular': request.POST.get('celular'),
            }
            ClienteDAO.actualizar(documento, datos)
            messages.success(request, 'Cliente actualizado exitosamente')
            return redirect('clientes_listar')
        except Exception as e:
            messages.error(request, f'Error al actualizar el cliente: {str(e)}')
    
    return render(request, 'clientes/editar.html', {'cliente': cliente})

@login_required
def clientes_eliminar(request, documento):
    try:
        ClienteDAO.eliminar(documento)
        messages.success(request, 'Cliente eliminado exitosamente')
    except Exception as e:
        messages.error(request, f'Error al eliminar el cliente: {str(e)}')
    return redirect('clientes_listar')

@login_required
def clientes_ver(request, documento):
    cliente = get_object_or_404(Cliente, documento=documento)
    historial = HistorialMembresia.objects.filter(cliente=cliente)
    pagos = PagoDAO.obtener_por_cliente(cliente)
    bonos = Bono.objects.filter(cliente=cliente).order_by('-fecha_otorgado')
    return render(request, 'clientes/ver.html', {'cliente': cliente, 'historial': historial, 'pagos': pagos, 'bonos': bonos})

@login_required
def clientes_renovar(request, documento):
    cliente = get_object_or_404(Cliente, documento=documento)
    
    if request.method == 'POST':
        try:
            membresia = get_object_or_404(Membresia, id=request.POST.get('membresia'))
            cliente.renovar_membresia(membresia)
            
            # AGREGAR BONO SI SE SELECCIONÓ (máximo 3 días)
            bono_dias = request.POST.get('bono_dias', '0')
            dias_extra = 0
            if bono_dias != '0':
                dias_extra = int(bono_dias)
                if dias_extra > 3:  # Validación
                    dias_extra = 3
                cliente.fecha_fin_membresia += timedelta(days=dias_extra)
                cliente.save()
                
                # REGISTRAR BONO
                tipo_bono_map = {
                    1: '1_dia',
                    2: '2_dias',
                    3: '3_dias',
                }
                
                Bono.objects.create(
                    cliente=cliente,
                    tipo_bono=tipo_bono_map.get(dias_extra, '1_dia'),
                    dias_regalo=dias_extra,
                    motivo='Regalo de renovación',
                    usuario_otorgo=request.user,
                    aplicado=True,
                    fecha_aplicado=timezone.now()
                )
            
            HistorialMembresia.objects.create(
                cliente=cliente,
                membresia=membresia,
                fecha_inicio=cliente.fecha_inicio_membresia,
                fecha_fin=cliente.fecha_fin_membresia,
                precio_pagado=membresia.precio
            )
            
            # Enviar email de renovación
            try:
                EmailService.enviar_email_renovacion(cliente)
                messages.success(request, 'Membresía renovada exitosamente y email enviado')
            except:
                messages.success(request, 'Membresía renovada exitosamente (no se pudo enviar el email)')
            
            return redirect('clientes_listar')
        except Exception as e:
            messages.error(request, f'Error al renovar la membresía: {str(e)}')
    
    membresias = MembresiaDAO.obtener_todas()
    return render(request, 'clientes/renovar.html', {'cliente': cliente, 'membresias': membresias})

@login_required
def clientes_importar_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        
        try:
            df = pd.read_excel(archivo)
            count = 0
            
            for index, row in df.iterrows():
                try:
                    membresia = Membresia.objects.get(id=int(row['membresia_id']))
                    fecha_inicio = timezone.now().date()
                    fecha_fin = fecha_inicio + timedelta(days=membresia.duracion_dias)
                    
                    cliente = Cliente.objects.create(
                        documento=str(row['documento']),
                        nombres=row['nombres'],
                        apellidos=row['apellidos'],
                        peso=float(row['peso']),
                        fecha_nacimiento=row['fecha_nacimiento'],
                        email=row['email'],
                        celular=str(row['celular']),
                        membresia_actual=membresia,
                        fecha_inicio_membresia=fecha_inicio,
                        fecha_fin_membresia=fecha_fin,
                        estado='activo'
                    )
                    
                    HistorialMembresia.objects.create(
                        cliente=cliente,
                        membresia=membresia,
                        fecha_inicio=fecha_inicio,
                        fecha_fin=fecha_fin,
                        precio_pagado=membresia.precio
                    )
                    count += 1
                except Exception as e:
                    continue
            
            messages.success(request, f'{count} clientes importados exitosamente')
        except Exception as e:
            messages.error(request, f'Error al importar: {str(e)}')
        
        return redirect('clientes_listar')
    
    return render(request, 'clientes/importar.html')

# ============= ASISTENCIAS =============
@login_required
def asistencias_listar(request):
    """Listar asistencias con filtros por mes/año"""
    
    # Obtener parámetros de filtro
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')  # Sin tilde, como en el template
    
    # Generar años disponibles (últimos 5 años + año actual)
    año_actual = timezone.now().year
    anios_disponibles = list(range(año_actual - 4, año_actual + 1))
    
    # Mes y año actuales para usar por defecto
    mes_actual = timezone.now().month
    
    # Si no hay filtros, usar mes y año actuales
    if not mes:
        mes = mes_actual
    else:
        mes = int(mes)
    
    if not anio:
        anio = año_actual
    else:
        anio = int(anio)
    
    # Filtrar asistencias por mes y año
    asistencias = Asistencia.objects.filter(
        fecha__year=anio,
        fecha__month=mes
    ).select_related('cliente', 'usuario_registro').order_by('-fecha')
    
    # Calcular estadísticas
    total_asistencias = asistencias.count()
    clientes_unicos = asistencias.values('cliente').distinct().count()
    
    # Obtener nombre del mes
    meses_nombres = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    mes_nombre = meses_nombres.get(mes, 'Mes')
    
    context = {
        'asistencias': asistencias,
        'anios_disponibles': anios_disponibles,  # Como lo espera el template
        'mes': mes,
        'anio': anio,  # Sin tilde
        'total_asistencias': total_asistencias,
        'clientes_unicos': clientes_unicos,
        'mes_nombre': mes_nombre,  # Faltaba esta variable
    }
    
    return render(request, 'asistencias/listar.html', context)

@login_required
def asistencias_registrar(request):
    if request.method == 'POST':
        documento = request.POST.get('documento')
        
        try:
            cliente = Cliente.objects.get(documento=documento)
            
            if cliente.estado != 'activo':
                return JsonResponse({'success': False, 'message': 'Cliente inactivo. Debe renovar membresía'})
            
            if cliente.fecha_fin_membresia < timezone.now().date():
                cliente.estado = 'inactivo'
                cliente.save()
                return JsonResponse({'success': False, 'message': 'Membresía vencida'})
            
            asistencia = Asistencia.objects.create(
                cliente=cliente,
                usuario_registro=request.user
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Asistencia registrada para {cliente.nombres} {cliente.apellidos}',
                'cliente': f'{cliente.nombres} {cliente.apellidos}',
                'hora': asistencia.hora.strftime('%H:%M:%S')
            })
        except Cliente.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Cliente no encontrado'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    # SOLUCIÓN: Usar date.today() directamente
    fecha_hoy = date.today()
    
    # Filtrar SOLO asistencias de HOY
    asistencias_hoy = Asistencia.objects.filter(
        fecha=fecha_hoy
    ).select_related('cliente').order_by('-hora')
    
    # DEBUG: Imprimir en consola para verificar
    print(f"Fecha de hoy: {fecha_hoy}")
    print(f"Total asistencias de hoy: {asistencias_hoy.count()}")
    for a in asistencias_hoy:
        print(f"  - {a.cliente.nombres}: {a.fecha} {a.hora}")
    
    context = {
        'asistencias_hoy': asistencias_hoy,
    }
    
    return render(request, 'asistencias/registrar.html', context)

@login_required
def asistencias_exportar_excel(request):
    fecha = request.GET.get('fecha', timezone.now().date())
    asistencias = Asistencia.objects.filter(fecha=fecha)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=asistencias_{fecha}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Asistencias'
    
    headers = ['Documento', 'Nombre', 'Hora', 'Usuario Registro']
    ws.append(headers)
    
    for asistencia in asistencias:
        ws.append([
            asistencia.cliente.documento,
            f"{asistencia.cliente.nombres} {asistencia.cliente.apellidos}",
            asistencia.hora.strftime('%H:%M:%S'),
            asistencia.usuario_registro.nombre if asistencia.usuario_registro else 'N/A'
        ])
    
    wb.save(response)
    return response

@login_required
def asistencias_exportar_pdf(request):
    fecha = request.GET.get('fecha', timezone.now().date())
    asistencias = Asistencia.objects.filter(fecha=fecha)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=asistencias_{fecha}.pdf'
    
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    
    p.setFont("Helvetica-Bold", 16)
    p.drawString(200, height - 50, f"Asistencias del {fecha}")
    
    y = height - 100
    p.setFont("Helvetica", 10)
    
    for asistencia in asistencias:
        texto = f"{asistencia.cliente.documento} - {asistencia.cliente.nombres} {asistencia.cliente.apellidos} - {asistencia.hora.strftime('%H:%M:%S')}"
        p.drawString(50, y, texto)
        y -= 20
        
        if y < 50:
            p.showPage()
            y = height - 50
    
    p.save()
    return response

# ============= PAGOS =============
@login_required
def pagos_listar(request):
    filtro = request.GET.get('filtro', 'todos')
    
    if filtro == 'pendientes':
        pagos = PagoDAO.obtener_pendientes()
    elif filtro == 'validados':
        pagos = PagoDAO.obtener_validados()
    else:
        pagos = PagoDAO.obtener_todos()
    
    return render(request, 'pagos/listar.html', {'pagos': pagos, 'filtro': filtro})

@login_required
def pagos_crear(request):
    if request.method == 'POST':
        documento = request.POST.get('documento')
        metodo_pago = request.POST.get('metodo_pago')
        tipo_pago = request.POST.get('tipo_pago', 'membresia')
        concepto = request.POST.get('concepto', '')
        observaciones = request.POST.get('observaciones', '')
        comprobante = request.POST.get('comprobante', '')
        
        try:
            # Buscar cliente por documento
            cliente = Cliente.objects.get(documento=documento)
            
            # Verificar que el cliente tenga membresía activa
            if not cliente.membresia_actual:
                messages.error(request, 'El cliente no tiene una membresía registrada.')
                return redirect('pagos_crear')
            
            # Obtener la membresía y el monto
            membresia = cliente.membresia_actual
            monto = membresia.precio
            
            # Si el concepto está vacío, usar el nombre de la membresía
            if not concepto:
                concepto = f'Pago de membresía {membresia.nombre}'
            
            # Crear el pago PRIMERO
            datos = {
                'cliente': cliente,
                'membresia': membresia,
                'concepto': concepto,
                'tipo_pago': tipo_pago,
                'monto': float(monto),
                'metodo_pago': metodo_pago,
                'comprobante': comprobante,
                'observaciones': observaciones,
                'usuario_registro': request.user,
                'estado': 'pendiente',
            }
            
            PagoDAO.crear(datos)
            
            # Actualizar fechas de membresía SI el pago es de tipo membresía
            if tipo_pago == 'membresia':
                if cliente.fecha_fin_membresia and cliente.fecha_fin_membresia > timezone.now().date():
                    # Si tiene membresía activa, sumar días
                    cliente.fecha_fin_membresia += timedelta(days=membresia.duracion_dias)
                else:
                    # Si no tiene o está vencida, iniciar desde hoy
                    cliente.fecha_inicio_membresia = timezone.now().date()
                    cliente.fecha_fin_membresia = timezone.now().date() + timedelta(days=membresia.duracion_dias)
            
            # IMPORTANTE: Cambiar estado del cliente a PENDIENTE
            cliente.estado = 'pendiente'
            cliente.save()
            
            messages.success(request, f'✓ Pago registrado exitosamente para {cliente.nombres} {cliente.apellidos}. Pendiente de validación.')
            return redirect('pagos_listar')
            
        except Cliente.DoesNotExist:
            messages.error(request, '✗ Cliente no encontrado. Verifique el documento.')
            return redirect('pagos_crear')
        except Exception as e:
            messages.error(request, f'✗ Error al registrar el pago: {str(e)}')
            return redirect('pagos_crear')
    
    # GET: Mostrar formulario
    metodos_pago = [
        ('efectivo', 'Efectivo'),
        ('tarjeta', 'Tarjeta'),
        ('transferencia', 'Transferencia'),
        ('nequi', 'Nequi'),
        ('daviplata', 'Daviplata'),
    ]
    
    context = {
        'metodos_pago': metodos_pago,
    }
    
    return render(request, 'pagos/crear.html', context)

@login_required
def api_buscar_cliente(request, documento):
    """API para buscar cliente por documento"""
    try:
        cliente = Cliente.objects.get(documento=documento)
        
        # Verificar que tenga membresía
        if not cliente.membresia_actual:
            return JsonResponse({
                'success': False,
                'message': 'El cliente no tiene una membresía registrada'
            })
        
        return JsonResponse({
            'success': True,
            'cliente': {
                'nombres': cliente.nombres,
                'apellidos': cliente.apellidos,
                'documento': cliente.documento,
                'membresia': cliente.membresia_actual.nombre,
                'monto': float(cliente.membresia_actual.precio),
                'estado': cliente.estado,
            }
        })
    except Cliente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Cliente no encontrado'
        })

@login_required
def pagos_registrar(request, documento):
    """Registrar pago de membresía desde la gestión de clientes"""
    from datetime import timedelta

    cliente = get_object_or_404(Cliente, documento=documento)
    membresias = MembresiaDAO.obtener_todas()

    if request.method == 'POST':
        try:
            membresia_id = request.POST.get('membresia_id')
            metodo_pago = request.POST.get('metodo_pago')
            monto = request.POST.get('monto')
            concepto = request.POST.get('concepto', 'Pago de membresía')
            referencia = request.POST.get('referencia', '')
            observaciones = request.POST.get('observaciones', '')

            membresia = get_object_or_404(Membresia, id=membresia_id)

            datos = {
                'cliente': cliente,
                'membresia': membresia,
                'concepto': concepto,
                'tipo_pago': 'membresia',
                'monto': float(monto),
                'metodo_pago': metodo_pago,
                'comprobante': referencia,
                'observaciones': observaciones,
                'usuario_registro': request.user,
                'estado': 'pendiente',  # SIEMPRE EN minúscula
            }

            PagoDAO.crear(datos)

            # Actualizar fechas de membresía
            if cliente.fecha_fin_membresia and cliente.fecha_fin_membresia > timezone.now().date():
                cliente.fecha_fin_membresia += timedelta(days=membresia.duracion_dias)
            else:
                cliente.fecha_inicio_membresia = timezone.now().date()
                cliente.fecha_fin_membresia = timezone.now().date() + timedelta(days=membresia.duracion_dias)

            cliente.estado = 'pendiente'
            cliente.save()

            messages.success(request, f'Pago registrado exitosamente. Pendiente de validación. Membresía válida hasta {cliente.fecha_fin_membresia.strftime("%d/%m/%Y")}')
            return redirect('clientes_listar')

        except Exception as e:
            messages.error(request, f'Error al registrar el pago: {str(e)}')

    context = {
        'cliente': cliente,
        'membresias': membresias,
    }

    return render(request, 'pagos/registrar.html', context)


@login_required
def pagos_ver(request, id):
    pago = get_object_or_404(Pago, id=id)
    return render(request, 'pagos/ver.html', {'pago': pago})

@login_required
def pagos_editar(request, id):
    pago = get_object_or_404(Pago, id=id)
    
    if pago.estado != 'pendiente':
        messages.error(request, 'Solo se pueden editar pagos pendientes')
        return redirect('pagos_listar')
    
    if request.method == 'POST':
        try:
            datos = {
                'concepto': request.POST.get('concepto'),
                'tipo_pago': request.POST.get('tipo_pago'),
                'monto': float(request.POST.get('monto')),
                'metodo_pago': request.POST.get('metodo_pago'),
                'comprobante': request.POST.get('comprobante', ''),
                'observaciones': request.POST.get('observaciones', ''),
            }
            
            if request.POST.get('membresia'):
                datos['membresia'] = get_object_or_404(Membresia, id=request.POST.get('membresia'))
            
            PagoDAO.actualizar(id, datos)
            messages.success(request, 'Pago actualizado exitosamente')
            return redirect('pagos_listar')
        except Exception as e:
            messages.error(request, f'Error al actualizar el pago: {str(e)}')
    
    membresias = MembresiaDAO.obtener_todas()
    return render(request, 'pagos/editar.html', {'pago': pago, 'membresias': membresias})

@login_required
@user_passes_test(es_administrador)
def pagos_validar(request, id):
    pago = get_object_or_404(Pago, id=id)
    
    if pago.estado != 'pendiente':
        messages.error(request, 'Este pago ya fue procesado')
        return redirect('pagos_listar')
    
    if request.method == 'POST':
        accion = request.POST.get('accion')
        
        if accion == 'validar':
            pago.validar_pago(request.user)
            messages.success(request, 'Pago validado exitosamente')

        elif accion == 'rechazar':
            observacion = request.POST.get('observacion_rechazo', 'Sin observación')
            pago.rechazar_pago(request.user, observacion)
            messages.success(request, 'Pago rechazado. Cliente marcado como inactivo.')
        
        return redirect('pagos_listar')
    
    return render(request, 'pagos/validar.html', {'pago': pago})

@login_required
def pagos_eliminar(request, id):
    try:
        PagoDAO.eliminar(id)
        messages.success(request, 'Pago eliminado exitosamente')
    except Exception as e:
        messages.error(request, f'Error al eliminar el pago: {str(e)}')
    return redirect('pagos_listar')

@login_required
def pagos_reportes(request):
    estadisticas = PagoDAO.obtener_estadisticas()
    
    pagos_reporte = []
    if request.GET.get('fecha_inicio') and request.GET.get('fecha_fin'):
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        pagos_reporte = PagoDAO.obtener_reporte_fechas(fecha_inicio, fecha_fin)
    
    context = {
        'estadisticas': estadisticas,
        'pagos_reporte': pagos_reporte,
        'fecha_inicio': request.GET.get('fecha_inicio', ''),
        'fecha_fin': request.GET.get('fecha_fin', ''),
    }
    
    return render(request, 'pagos/reportes.html', context)

@login_required
def pagos_exportar_excel(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio and fecha_fin:
        pagos = PagoDAO.obtener_reporte_fechas(fecha_inicio, fecha_fin)
    else:
        pagos = PagoDAO.obtener_validados()
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=reporte_pagos_{timezone.now().date()}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pagos'
    
    headers = ['ID', 'Cliente', 'Concepto', 'Tipo', 'Monto', 'Método', 'Estado', 'Fecha Pago', 'Usuario Registro']
    ws.append(headers)
    
    for pago in pagos:
        ws.append([
            pago.id,
            f"{pago.cliente.nombres} {pago.cliente.apellidos}",
            pago.concepto,
            pago.get_tipo_pago_display(),
            float(pago.monto),
            pago.get_metodo_pago_display(),
            pago.get_estado_display(),
            pago.fecha_pago.strftime('%d/%m/%Y %H:%M'),
            pago.usuario_registro.nombre if pago.usuario_registro else 'N/A'
        ])
    
    wb.save(response)
    return response

@login_required
def pagos_exportar_pdf(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio and fecha_fin:
        pagos = PagoDAO.obtener_reporte_fechas(fecha_inicio, fecha_fin)
        titulo = f"Reporte de Pagos del {fecha_inicio} al {fecha_fin}"
    else:
        pagos = PagoDAO.obtener_validados()[:50]
        titulo = "Reporte de Pagos Validados"
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=reporte_pagos_{timezone.now().date()}.pdf'
    
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    
    p.setFont("Helvetica-Bold", 16)
    p.drawString(150, height - 50, titulo)
    
    y = height - 100
    p.setFont("Helvetica", 9)
    
    total = 0
    for pago in pagos:
        texto = f"#{pago.id} - {pago.cliente.nombres} {pago.cliente.apellidos} - ${pago.monto} - {pago.get_metodo_pago_display()}"
        p.drawString(50, y, texto)
        total += float(pago.monto)
        y -= 15
        
        if y < 50:
            p.showPage()
            y = height - 50
            p.setFont("Helvetica", 9)
    
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y - 20, f"TOTAL: ${total:,.2f} COP")
    
    p.save()
    return response

@login_required
def pagos_registrar(request, documento):
    """Registrar pago de membresía para un cliente"""
    cliente = get_object_or_404(Cliente, documento=documento)
    membresias = Membresia.objects.filter(activa=True).order_by('precio')
    
    if request.method == 'POST':
        membresia_id = request.POST.get('membresia_id')
        metodo_pago = request.POST.get('metodo_pago')
        monto = request.POST.get('monto')
        
        try:
            membresia = Membresia.objects.get(id=membresia_id)
            
            # Crear el pago
            pago = Pago.objects.create(
                cliente=cliente,
                membresia=membresia,
                monto=monto,
                metodo_pago=metodo_pago,
                fecha_pago=timezone.now(),
                usuario_registro=request.user
            )
            
            # Actualizar la membresía del cliente
            if cliente.fecha_fin_membresia and cliente.fecha_fin_membresia > timezone.now().date():
                cliente.fecha_fin_membresia += timedelta(days=membresia.duracion_dias)
            else:
                cliente.fecha_inicio_membresia = timezone.now().date()
                cliente.fecha_fin_membresia = timezone.now().date() + timedelta(days=membresia.duracion_dias)
            
            cliente.estado = 'Pendiente'
            cliente.save()
            
            messages.success(request, f'Pago registrado exitosamente. Membresía válida hasta {cliente.fecha_fin_membresia.strftime("%d/%m/%Y")}')
            return redirect('clientes_listar')
            
        except Exception as e:
            messages.error(request, f'Error al registrar el pago: {str(e)}')
    
    context = {
        'cliente': cliente,
        'membresias': membresias,
    }
    
    return render(request, 'pagos/registrar.html', context)


# ============= USUARIOS =============
@login_required
@user_passes_test(es_administrador)
def usuarios_listar(request):
    usuarios = UsuarioDAO.obtener_todos()
    return render(request, 'usuarios/listar.html', {'usuarios': usuarios})

@login_required
@user_passes_test(es_administrador)
def usuarios_crear(request):
    if request.method == 'POST':
        try:
            datos = {
                'nombre': request.POST.get('nombre'),
                'correo': request.POST.get('correo'),
                'password': request.POST.get('password'),
                'rol': request.POST.get('rol'),
            }
            UsuarioDAO.crear(datos)
            messages.success(request, 'Usuario creado exitosamente')
            return redirect('usuarios_listar')
        except Exception as e:
            messages.error(request, f'Error al crear el usuario: {str(e)}')
    
    return render(request, 'usuarios/crear.html')

@login_required
@user_passes_test(es_administrador)
def usuarios_editar(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    
    if request.method == 'POST':
        try:
            datos = {
                'nombre': request.POST.get('nombre'),
                'correo': request.POST.get('correo'),
                'rol': request.POST.get('rol'),
            }
            
            password = request.POST.get('password')
            if password:
                datos['password'] = password
            
            UsuarioDAO.actualizar(id, datos)
            messages.success(request, 'Usuario actualizado exitosamente')
            return redirect('usuarios_listar')
        except Exception as e:
            messages.error(request, f'Error al actualizar el usuario: {str(e)}')
    
    return render(request, 'usuarios/editar.html', {'usuario': usuario})

@login_required
@user_passes_test(es_administrador)
def usuarios_eliminar(request, id):
    try:
        UsuarioDAO.eliminar(id)
        messages.success(request, 'Usuario eliminado exitosamente')
    except Exception as e:
        messages.error(request, f'Error al eliminar el usuario: {str(e)}')
    return redirect('usuarios_listar')

@login_required
@user_passes_test(es_administrador)
def usuarios_ver(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    return render(request, 'usuarios/ver.html', {'usuario': usuario})

# ============= BONOS Y REGALOS (1-3 DÍAS) =============

@login_required
@user_passes_test(es_administrador)
def bonos_listar(request):
    """Listar todos los bonos otorgados"""
    bonos = Bono.objects.all().select_related('cliente', 'usuario_otorgo')
    
    # Filtros opcionales
    filtro_estado = request.GET.get('estado', 'todos')
    if filtro_estado == 'pendientes':
        bonos = bonos.filter(aplicado=False)
    elif filtro_estado == 'aplicados':
        bonos = bonos.filter(aplicado=True)
    
    context = {
        'bonos': bonos,
        'filtro_estado': filtro_estado,
        'total_bonos': bonos.count(),
    }
    return render(request, 'bonos/listar.html', context)

@login_required
@user_passes_test(es_administrador)
def bonos_crear(request):
    """Crear un nuevo bono/regalo para un cliente (1-3 días)"""
    if request.method == 'POST':
        try:
            cliente = get_object_or_404(Cliente, documento=request.POST.get('cliente'))
            tipo_bono = request.POST.get('tipo_bono')
            
            # Mapear tipo de bono a días (máximo 3)
            dias_map = {
                '1_dia': 1,
                '2_dias': 2,
                '3_dias': 3,
            }
            
            bono = Bono.objects.create(
                cliente=cliente,
                tipo_bono=tipo_bono,
                dias_regalo=dias_map.get(tipo_bono, 1),
                motivo=request.POST.get('motivo'),
                usuario_otorgo=request.user
            )
            
            # Aplicar automáticamente si el checkbox está marcado
            if request.POST.get('aplicar_ahora'):
                bono.aplicar_bono()
                messages.success(request, f'Bono de {bono.dias_regalo} días creado y aplicado a {cliente.nombres} {cliente.apellidos}')
            else:
                messages.success(request, f'Bono de {bono.dias_regalo} días creado para {cliente.nombres} {cliente.apellidos}')
            
            return redirect('bonos_listar')
        except Exception as e:
            messages.error(request, f'Error al crear el bono: {str(e)}')
    
    clientes = Cliente.objects.filter(estado='activo').order_by('nombres')
    return render(request, 'bonos/crear.html', {'clientes': clientes})

@login_required
@user_passes_test(es_administrador)
def bonos_aplicar(request, id):
    """Aplicar un bono pendiente"""
    bono = get_object_or_404(Bono, id=id)
    
    if bono.aplicado:
        messages.warning(request, 'Este bono ya fue aplicado')
    else:
        if bono.aplicar_bono():
            messages.success(request, f'Bono aplicado exitosamente. Se agregaron {bono.dias_regalo} días a la membresía')
        else:
            messages.error(request, 'Error al aplicar el bono. Verifica que el cliente tenga una membresía activa')
    
    return redirect('bonos_listar')

@login_required
@user_passes_test(es_administrador)
def bonos_eliminar(request, id):
    """Eliminar un bono (solo si no ha sido aplicado)"""
    try:
        bono = get_object_or_404(Bono, id=id)
        if bono.aplicado:
            messages.error(request, 'No se puede eliminar un bono ya aplicado')
        else:
            bono.delete()
            messages.success(request, 'Bono eliminado exitosamente')
    except Exception as e:
        messages.error(request, f'Error al eliminar el bono: {str(e)}')
    
    return redirect('bonos_listar')

@login_required
@user_passes_test(es_administrador)
def bonos_estadisticas(request):
    """Estadísticas de bonos otorgados"""
    total_bonos = Bono.objects.count()
    bonos_aplicados = Bono.objects.filter(aplicado=True).count()
    bonos_pendientes = Bono.objects.filter(aplicado=False).count()
    
    # Total de días regalados
    total_dias_regalados = Bono.objects.filter(aplicado=True).aggregate(
        total=Sum('dias_regalo')
    )['total'] or 0
    
    # Bonos por tipo
    bonos_por_tipo = Bono.objects.values('tipo_bono').annotate(
        total=Count('id')
    ).order_by('-total')
    
    context = {
        'total_bonos': total_bonos,
        'bonos_aplicados': bonos_aplicados,
        'bonos_pendientes': bonos_pendientes,
        'total_dias_regalados': total_dias_regalados,
        'bonos_por_tipo': bonos_por_tipo,
    }
    
    return render(request, 'bonos/estadisticas.html', context)

# ============= COLORES Y UTILIDADES EXCEL =============

# Colores corporativos FITTECH
COLOR_HEADER = "667EEA"
COLOR_SUBHEADER = "764BA2"
COLOR_ALT_ROW = "F3F4F6"
COLOR_TOTAL = "10B981"

def aplicar_estilos_header(ws, row, columnas):
    """Aplicar estilos al encabezado"""
    font_header = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, valor in enumerate(columnas, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = valor
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = alignment
        cell.border = border
    
    ws.row_dimensions[row].height = 25

def aplicar_estilos_fila(ws, row, valores, alternado=False):
    """Aplicar estilos a una fila de datos"""
    font = Font(name='Calibri', size=11)
    fill = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type='solid') if alternado else PatternFill()
    alignment = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    for col, valor in enumerate(valores, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = valor
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border
        
        if isinstance(valor, (int, float)) and not isinstance(valor, bool):
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if '.' in str(valor) or isinstance(valor, float):
                cell.number_format = '$#,##0.00'

def agregar_titulo_excel(ws, titulo, subtitulo=""):
    """Agregar título al Excel"""
    ws.merge_cells('A1:F1')
    cell_titulo = ws['A1']
    cell_titulo.value = titulo
    cell_titulo.font = Font(name='Calibri', size=16, bold=True, color='667EEA')
    cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:F2')
    cell_sub = ws['A2']
    cell_sub.value = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}" + (f" | {subtitulo}" if subtitulo else "")
    cell_sub.font = Font(name='Calibri', size=10, italic=True, color='6B7280')
    cell_sub.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18
    
    return 4

def ajustar_ancho_columnas(ws):
    """Ajustar ancho automático de columnas"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============= REPORTES GENERALES (MANTENER ORIGINAL) =============

@login_required
def reportes_generales(request):
    """Vista principal de reportes con estadísticas generales - CORREGIDA"""
    
    # Estadísticas Membresías
    total_membresias = Membresia.objects.filter(activa=True).count()
    membresias_stats = Membresia.objects.filter(activa=True).aggregate(
        total_ingresos_potenciales=models.Sum('precio'),
        precio_promedio=models.Avg('precio')
    )
    
    # Estadísticas Clientes
    total_clientes = Cliente.objects.count()
    clientes_activos = Cliente.objects.filter(estado='activo').count()
    clientes_inactivos = Cliente.objects.filter(estado='inactivo').count()
    
    # Estadísticas Asistencias
    ahora = timezone.now()
    hoy = ahora.date()
    inicio_mes = hoy.replace(day=1)
    
    asistencias_hoy = Asistencia.objects.filter(fecha=hoy).count()
    asistencias_mes = Asistencia.objects.filter(fecha__gte=inicio_mes).count()
    
    # CORRECCIÓN: Estadísticas Pagos usando rangos de datetime
    inicio_dia = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
    fin_dia = ahora.replace(hour=23, minute=59, second=59, microsecond=999999)
    inicio_mes_dt = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Pagos con rangos de datetime
    pagos_stats = {
        'total_pagos': Pago.objects.filter(estado='validado').count(),
        'total_ingresos': float(Pago.objects.filter(estado='validado').aggregate(Sum('monto'))['monto__sum'] or 0),
        'pagos_pendientes': Pago.objects.filter(estado='pendiente').count(),
        'pagos_rechazados': Pago.objects.filter(estado='rechazado').count(),
        'pagos_hoy': Pago.objects.filter(
            fecha_pago__gte=inicio_dia,
            fecha_pago__lte=fin_dia,
            estado='validado'
        ).count(),
        'ingresos_hoy': float(Pago.objects.filter(
            fecha_pago__gte=inicio_dia,
            fecha_pago__lte=fin_dia,
            estado='validado'
        ).aggregate(Sum('monto'))['monto__sum'] or 0),
        'ingresos_mes': float(Pago.objects.filter(
            fecha_pago__gte=inicio_mes_dt,
            fecha_pago__lte=ahora,
            estado='validado'
        ).aggregate(Sum('monto'))['monto__sum'] or 0),
    }
    
    # Pagos por método
    pagos_stats['pagos_por_metodo'] = Pago.objects.filter(estado='validado').values('metodo_pago').annotate(
        total=Count('id'),
        monto_total=Sum('monto')
    ).order_by('-monto_total')
    
    # Estadísticas Usuarios
    total_usuarios = Usuario.objects.count()
    usuarios_admin = Usuario.objects.filter(rol='administrador').count()
    usuarios_empleados = Usuario.objects.filter(rol='empleado').count()
    
    context = {
        'total_membresias': total_membresias,
        'membresias_stats': membresias_stats,
        'total_clientes': total_clientes,
        'clientes_activos': clientes_activos,
        'clientes_inactivos': clientes_inactivos,
        'asistencias_hoy': asistencias_hoy,
        'asistencias_mes': asistencias_mes,
        'pagos_stats': pagos_stats,
        'total_usuarios': total_usuarios,
        'usuarios_admin': usuarios_admin,
        'usuarios_empleados': usuarios_empleados,
    }
    
    return render(request, 'reportes/generales.html', context)

# ============= REPORTES CLIENTES =============

@login_required
def reportes_clientes_excel(request):
    """Generar reporte de clientes en Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    fila = agregar_titulo_excel(ws, "REPORTE DE CLIENTES FITTECH")
    
    columnas = ['Documento', 'Nombres', 'Apellidos', 'Email', 'Teléfono', 'Membresía', 'Estado', 'Fecha Fin']
    aplicar_estilos_header(ws, fila, columnas)
    fila += 1
    
    clientes = Cliente.objects.select_related('membresia_actual').all()
    
    for idx, cliente in enumerate(clientes):
        valores = [
            cliente.documento,
            cliente.nombres,
            cliente.apellidos,
            cliente.email,
            cliente.celular,
            cliente.membresia_actual.nombre if cliente.membresia_actual else 'N/A',
            cliente.get_estado_display().upper(),
            cliente.fecha_fin_membresia.strftime('%d/%m/%Y') if cliente.fecha_fin_membresia else 'N/A'
        ]
        aplicar_estilos_fila(ws, fila, valores, alternado=(idx % 2 == 0))
        fila += 1
    
    fila += 1
    ws.merge_cells(f'A{fila}:D{fila}')
    cell = ws[f'A{fila}']
    cell.value = f"TOTAL CLIENTES: {clientes.count()}"
    cell.font = Font(bold=True, size=12, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    ajustar_ancho_columnas(ws)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Clientes_{timezone.now().strftime("%d_%m_%Y")}.xlsx"'
    wb.save(response)
    return response

@login_required
def reportes_clientes_pdf(request):
    """Generar reporte de clientes en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    doc.title = "Reporte de Clientes FITTECH"
    doc.author = "FITTECH - Sistema de Gestión"
    doc.subject = "Reporte de Clientes"

    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#667EEA'),
        spaceAfter=10,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("REPORTE DE CLIENTES FITTECH", title_style))
    elements.append(Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    clientes = Cliente.objects.select_related('membresia_actual').all()
    
    data = [['Documento', 'Nombre', 'Email', 'Teléfono', 'Membresía', 'Estado']]
    
    for cliente in clientes:
        data.append([
            cliente.documento,
            f"{cliente.nombres} {cliente.apellidos}",
            cliente.email,
            cliente.celular,
            cliente.membresia_actual.nombre if cliente.membresia_actual else 'N/A',
            cliente.get_estado_display()
        ])
    
    table = Table(data, colWidths=[80, 120, 120, 80, 80, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"<b>Total Clientes: {clientes.count()}</b>", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return HttpResponse(buffer, content_type='application/pdf')

# ============= REPORTES MEMBRESÍAS =============

@login_required
@user_passes_test(es_administrador)
def reportes_membresias_excel(request):
    """Generar reporte de membresías en Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Membresías"
    
    fila = agregar_titulo_excel(ws, "REPORTE DE MEMBRESÍAS FITTECH")
    
    columnas = ['Nombre', 'Duración (días)', 'Precio', 'Descripción', 'Estado', 'Clientes Activos']
    aplicar_estilos_header(ws, fila, columnas)
    fila += 1
    
    membresias = Membresia.objects.all()
    
    for idx, membresia in enumerate(membresias):
        clientes_activos = Cliente.objects.filter(membresia_actual=membresia, estado='activo').count()
        valores = [
            membresia.nombre,
            membresia.duracion_dias,
            float(membresia.precio),
            membresia.descripcion or 'N/A',
            'ACTIVA' if membresia.activa else 'INACTIVA',
            clientes_activos
        ]
        aplicar_estilos_fila(ws, fila, valores, alternado=(idx % 2 == 0))
        fila += 1
    
    fila += 1
    ws.merge_cells(f'A{fila}:D{fila}')
    cell = ws[f'A{fila}']
    cell.value = f"TOTAL MEMBRESÍAS: {membresias.count()}"
    cell.font = Font(bold=True, size=12, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    ajustar_ancho_columnas(ws)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Membresias_{timezone.now().strftime("%d_%m_%Y")}.xlsx"'
    wb.save(response)
    return response

@login_required
@user_passes_test(es_administrador)
def reportes_membresias_pdf(request):
    """Generar reporte de membresías en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    doc.title = "Reporte de Membresías FITTECH"
    doc.author = "FITTECH - Sistema de Gestión"
    doc.subject = "Reporte de Membresías"

    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#667EEA'),
        spaceAfter=10,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("REPORTE DE MEMBRESÍAS FITTECH", title_style))
    elements.append(Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    membresias = Membresia.objects.all()
    
    data = [['Nombre', 'Duración', 'Precio', 'Estado', 'Clientes']]
    
    for membresia in membresias:
        clientes_activos = Cliente.objects.filter(membresia_actual=membresia, estado='activo').count()
        data.append([
            membresia.nombre,
            f"{membresia.duracion_dias} días",
            f"${membresia.precio:,.0f}",
            'Activa' if membresia.activa else 'Inactiva',
            str(clientes_activos)
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"<b>Total Membresías: {membresias.count()}</b>", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return HttpResponse(buffer, content_type='application/pdf')

# ============= REPORTES USUARIOS =============

@login_required
@user_passes_test(es_administrador)
def reportes_usuarios_excel(request):
    """Generar reporte de usuarios en Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    fila = agregar_titulo_excel(ws, "REPORTE DE USUARIOS FITTECH")
    
    columnas = ['Nombre', 'Correo', 'Rol', 'Estado', 'Fecha Creación']
    aplicar_estilos_header(ws, fila, columnas)
    fila += 1
    
    usuarios = Usuario.objects.all()
    
    for idx, usuario in enumerate(usuarios):
        valores = [
            usuario.nombre,
            usuario.correo,
            usuario.get_rol_display().upper(),
            'ACTIVO' if usuario.is_active else 'INACTIVO',
            usuario.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        ]
        aplicar_estilos_fila(ws, fila, valores, alternado=(idx % 2 == 0))
        fila += 1
    
    fila += 1
    ws.merge_cells(f'A{fila}:C{fila}')
    cell = ws[f'A{fila}']
    cell.value = f"TOTAL USUARIOS: {usuarios.count()}"
    cell.font = Font(bold=True, size=12, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    ajustar_ancho_columnas(ws)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Usuarios_{timezone.now().strftime("%d_%m_%Y")}.xlsx"'
    wb.save(response)
    return response

@login_required
@user_passes_test(es_administrador)
def reportes_usuarios_pdf(request):
    """Generar reporte de usuarios en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#667EEA'),
        spaceAfter=10,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("REPORTE DE USUARIOS FITTECH", title_style))
    elements.append(Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    usuarios = Usuario.objects.all()
    
    data = [['Nombre', 'Correo', 'Rol', 'Estado']]
    
    for usuario in usuarios:
        data.append([
            usuario.nombre,
            usuario.correo,
            usuario.get_rol_display(),
            'Activo' if usuario.is_active else 'Inactivo'
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"<b>Total Usuarios: {usuarios.count()}</b>", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return HttpResponse(buffer, content_type='application/pdf')

# ============= REPORTES DE PAGOS =============

@login_required
def pagos_exportar_excel(request):
    """Generar reporte de pagos en Excel con estilos profesionales"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"
    
    fila = agregar_titulo_excel(ws, "REPORTE DE PAGOS E INGRESOS FITTECH")
    
    # Encabezados
    columnas = ['ID', 'Cliente', 'Concepto', 'Monto', 'Método', 'Tipo', 'Estado', 'Fecha Pago', 'Usuario Registro']
    aplicar_estilos_header(ws, fila, columnas)
    fila += 1
    
    # Obtener filtros
    filtro_estado = request.GET.get('estado', 'todos')
    
    # Datos
    pagos = Pago.objects.select_related('cliente', 'usuario_registro').all()
    
    if filtro_estado != 'todos':
        pagos = pagos.filter(estado=filtro_estado)
    
    total_ingresos = 0
    pagos_validados = 0
    
    for idx, pago in enumerate(pagos.order_by('-fecha_pago')):
        valores = [
            pago.id,
            f"{pago.cliente.nombres} {pago.cliente.apellidos}",
            pago.concepto,
            float(pago.monto),
            pago.get_metodo_pago_display(),
            pago.tipo_pago.upper(),
            pago.get_estado_display().upper(),
            pago.fecha_pago.strftime('%d/%m/%Y %H:%M'),
            pago.usuario_registro.nombre
        ]
        aplicar_estilos_fila(ws, fila, valores, alternado=(idx % 2 == 0))
        
        if pago.estado == 'validado':
            total_ingresos += float(pago.monto)
            pagos_validados += 1
        
        fila += 1
    
    # Resumen
    fila += 1
    ws.merge_cells(f'A{fila}:D{fila}')
    cell = ws[f'A{fila}']
    cell.value = f"TOTAL PAGOS: {pagos.count()}"
    cell.font = Font(bold=True, size=12, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    fila += 1
    ws.merge_cells(f'A{fila}:D{fila}')
    cell = ws[f'A{fila}']
    cell.value = f"PAGOS VALIDADOS: {pagos_validados}"
    cell.font = Font(bold=True, size=12, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    fila += 1
    ws.merge_cells(f'A{fila}:D{fila}')
    cell = ws[f'A{fila}']
    cell.value = f"TOTAL INGRESOS: ${total_ingresos:,.2f} COP"
    cell.font = Font(bold=True, size=12, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    ajustar_ancho_columnas(ws)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Pagos_{timezone.now().strftime("%d_%m_%Y")}.xlsx"'
    wb.save(response)
    return response

@login_required
def pagos_exportar_pdf(request):
    """Generar reporte de pagos en PDF con estilos profesionales"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    doc.title = "Reporte de Pagos e Ingresos FITTECH"
    doc.author = "FITTECH - Sistema de Gestión"
    doc.subject = "Reporte de Pagos e Ingresos"

    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#667EEA'),
        spaceAfter=10,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("REPORTE DE PAGOS E INGRESOS FITTECH", title_style))
    elements.append(Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Obtener filtros
    filtro_estado = request.GET.get('estado', 'todos')
    
    pagos = Pago.objects.select_related('cliente', 'usuario_registro').all()
    
    if filtro_estado != 'todos':
        pagos = pagos.filter(estado=filtro_estado)
    
    pagos = pagos.order_by('-fecha_pago')[:50]  # Últimos 50
    
    data = [['ID', 'Cliente', 'Concepto', 'Monto', 'Método', 'Estado', 'Fecha']]
    
    total_ingresos = 0
    pagos_validados = 0
    
    for pago in pagos:
        data.append([
            str(pago.id),
            f"{pago.cliente.nombres} {pago.cliente.apellidos}",
            pago.concepto[:20] + '...' if len(pago.concepto) > 20 else pago.concepto,
            f"${pago.monto:,.0f}",
            pago.get_metodo_pago_display()[:10],
            pago.get_estado_display(),
            pago.fecha_pago.strftime('%d/%m/%Y')
        ])
        
        if pago.estado == 'validado':
            total_ingresos += float(pago.monto)
            pagos_validados += 1
    
    table = Table(data, colWidths=[30, 100, 80, 60, 60, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8)
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"<b>Total Pagos: {pagos.count()}</b>", styles['Normal']))
    elements.append(Paragraph(f"<b>Pagos Validados: {pagos_validados}</b>", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Ingresos: ${total_ingresos:,.2f} COP</b>", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return HttpResponse(buffer, content_type='application/pdf')

# ============= REPORTES DE ASISTENCIAS =============

@login_required
def asistencias_exportar_excel(request):
    """Generar reporte de asistencias en Excel con estilos profesionales"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"
    
    # Obtener filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    subtitulo = ""
    if fecha_inicio and fecha_fin:
        subtitulo = f"Del {fecha_inicio} al {fecha_fin}"
    elif not fecha_inicio and not fecha_fin:
        subtitulo = f"Hoy: {timezone.now().strftime('%d/%m/%Y')}"
    
    fila = agregar_titulo_excel(ws, "REPORTE DE ASISTENCIAS FITTECH", subtitulo)
    
    # Encabezados
    columnas = ['ID', 'Cliente', 'Documento', 'Teléfono', 'Membresía', 'Fecha', 'Hora', 'Usuario Registro']
    aplicar_estilos_header(ws, fila, columnas)
    fila += 1
    
    # Datos
    asistencias = Asistencia.objects.select_related('cliente', 'usuario_registro').all()
    
    if fecha_inicio:
        asistencias = asistencias.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        asistencias = asistencias.filter(fecha__lte=fecha_fin)
    
    # Si no hay filtros, mostrar solo de hoy
    if not fecha_inicio and not fecha_fin:
        asistencias = asistencias.filter(fecha=timezone.now().date())
    
    for idx, asistencia in enumerate(asistencias.order_by('-fecha', '-hora')):
        valores = [
            asistencia.id,
            f"{asistencia.cliente.nombres} {asistencia.cliente.apellidos}",
            asistencia.cliente.documento,
            asistencia.cliente.celular,
            asistencia.cliente.membresia_actual.nombre if asistencia.cliente.membresia_actual else 'N/A',
            asistencia.fecha.strftime('%d/%m/%Y'),
            asistencia.hora.strftime('%H:%M:%S') if asistencia.hora else 'N/A',
            asistencia.usuario_registro.nombre
        ]
        aplicar_estilos_fila(ws, fila, valores, alternado=(idx % 2 == 0))
        fila += 1
    
    # Resumen
    fila += 1
    ws.merge_cells(f'A{fila}:E{fila}')
    cell = ws[f'A{fila}']
    cell.value = f"TOTAL ASISTENCIAS: {asistencias.count()}"
    cell.font = Font(bold=True, size=12, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    ajustar_ancho_columnas(ws)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Asistencias_{timezone.now().strftime("%d_%m_%Y")}.xlsx"'
    wb.save(response)
    return response

@login_required
def asistencias_exportar_pdf(request):
    """Generar reporte de asistencias en PDF con estilos profesionales"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    doc.title = "Reporte de Asistencias FITTECH"
    doc.author = "FITTECH - Sistema de Gestión"
    doc.subject = "Reporte de Asistencias"

    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#667EEA'),
        spaceAfter=10,
        alignment=TA_CENTER
    )
    
    # Obtener filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    elements.append(Paragraph("REPORTE DE ASISTENCIAS FITTECH", title_style))
    elements.append(Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Datos
    asistencias = Asistencia.objects.select_related('cliente', 'usuario_registro').all()
    
    if fecha_inicio:
        asistencias = asistencias.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        asistencias = asistencias.filter(fecha__lte=fecha_fin)
    
    # Si no hay filtros, mostrar solo de hoy
    if not fecha_inicio and not fecha_fin:
        asistencias = asistencias.filter(fecha=timezone.now().date())
    
    asistencias = asistencias.order_by('-fecha', '-hora')[:50]  # Últimas 50
    
    data = [['Cliente', 'Documento', 'Teléfono', 'Fecha', 'Hora', 'Usuario']]
    
    for asistencia in asistencias:
        data.append([
            f"{asistencia.cliente.nombres} {asistencia.cliente.apellidos}",
            asistencia.cliente.documento,
            asistencia.cliente.celular,
            asistencia.fecha.strftime('%d/%m/%Y'),
            asistencia.hora.strftime('%H:%M') if asistencia.hora else 'N/A',
            asistencia.usuario_registro.nombre[:15]
        ])
    
    table = Table(data, colWidths=[100, 70, 70, 70, 50, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8)
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"<b>Total Asistencias: {asistencias.count()}</b>", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return HttpResponse(buffer, content_type='application/pdf')

# ============= REPORTE CONSOLIDADO =============

@login_required
@user_passes_test(es_administrador)
def reporte_consolidado_excel(request):
    """Generar reporte consolidado en Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"
    
    fila = agregar_titulo_excel(ws, "REPORTE CONSOLIDADO FITTECH")
    
    # Estadísticas
    total_clientes = Cliente.objects.count()
    clientes_activos = Cliente.objects.filter(estado='activo').count()
    clientes_inactivos = Cliente.objects.filter(estado='inactivo').count()
    total_membresias = Membresia.objects.filter(activa=True).count()
    total_asistencias = Asistencia.objects.count()
    asistencias_hoy = Asistencia.objects.filter(fecha=timezone.now().date()).count()
    total_pagos = Pago.objects.count()
    ingresos_total = float(Pago.objects.filter(estado='validado').aggregate(Sum('monto'))['monto__sum'] or 0)
    pagos_pendientes = Pago.objects.filter(estado='pendiente').count()
    total_usuarios = Usuario.objects.count()
    administradores = Usuario.objects.filter(rol='administrador').count()
    empleados = Usuario.objects.filter(rol='empleado').count()
    
    datos = [
        ['MEMBRESÍAS', '', ''],
        ['Total Membresías', total_membresias, ''],
        ['', '', ''],
        ['CLIENTES', '', ''],
        ['Total Clientes', total_clientes, ''],
        ['Activos', clientes_activos, ''],
        ['Inactivos', clientes_inactivos, ''],
        ['', '', ''],
        ['ASISTENCIAS', '', ''],
        ['Hoy', asistencias_hoy, ''],
        ['Total', total_asistencias, ''],
        ['', '', ''],
        ['PAGOS', '', ''],
        ['Total Pagos', total_pagos, ''],
        ['Ingresos Total', f"${ingresos_total:,.2f}", ''],
        ['Pendientes', pagos_pendientes, ''],
        ['', '', ''],
        ['USUARIOS', '', ''],
        ['Total Usuarios', total_usuarios, ''],
        ['Administradores', administradores, ''],
        ['Empleados', empleados, ''],
    ]
    
    for idx, fila_datos in enumerate(datos):
        if fila_datos[0] and not fila_datos[1]:
            ws.merge_cells(f'A{fila}:C{fila}')
            cell = ws[f'A{fila}']
            cell.value = fila_datos[0]
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            aplicar_estilos_fila(ws, fila, fila_datos, alternado=(idx % 2 == 0))
        
        fila += 1
    
    ajustar_ancho_columnas(ws)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Consolidado_{timezone.now().strftime("%d_%m_%Y")}.xlsx"'
    wb.save(response)
    return response



# ============= GESTIÓN DE EMAILS =============

@login_required
@user_passes_test(es_administrador)
def emails_panel(request):
    """Panel de control de emails - Vista principal"""
    from django.conf import settings
    
    dias_aviso = getattr(settings, 'DIAS_AVISO_VENCIMIENTO', 7)
    fecha_limite = timezone.now().date() + timedelta(days=dias_aviso)
    
    # Clientes con membresía por vencer
    clientes_por_vencer = Cliente.objects.filter(
        fecha_fin_membresia__lte=fecha_limite,
        fecha_fin_membresia__gte=timezone.now().date(),
        estado='activo'
    )
    
    # Clientes inactivos
    clientes_inactivos = Cliente.objects.filter(
        estado='inactivo'
    ).exclude(email__isnull=True).exclude(email='')
    
    context = {
        'clientes_por_vencer_lista': clientes_por_vencer,
        'total_por_vencer': clientes_por_vencer.count(),
        'total_inactivos': clientes_inactivos.count(),
        'dias_aviso': dias_aviso,
    }
    
    return render(request, 'emails/panel.html', context)


# ========== EMAILS PARA CLIENTES CON MEMBRESÍA POR VENCER ==========

@login_required
@user_passes_test(es_administrador)
def enviar_emails_vencimiento(request):
    """Enviar emails masivos a clientes con membresía por vencer"""
    if request.method == 'POST':
        resultado = EmailService.enviar_emails_masivos_vencimiento()
        
        if resultado['enviados'] > 0:
            messages.success(request, f"✓ Se enviaron {resultado['enviados']} emails correctamente")
        
        if resultado['fallidos'] > 0:
            messages.warning(request, f"⚠ {resultado['fallidos']} emails fallaron")
        
        if resultado['enviados'] == 0 and resultado['fallidos'] == 0:
            messages.info(request, 'No hay clientes con membresías por vencer')
        
        return redirect('emails_panel')
    
    return redirect('emails_panel')


@login_required
def enviar_email_vencimiento_individual(request, documento):
    """Enviar email de vencimiento a un cliente específico"""
    cliente = get_object_or_404(Cliente, documento=documento)
    
    if EmailService.enviar_email_vencimiento(cliente):
        messages.success(request, f'✓ Email de vencimiento enviado a {cliente.email}')
    else:
        messages.error(request, '✗ Error al enviar el email')
    
    return redirect('clientes_ver', documento=documento)

@login_required
def enviar_email_reactivacion_individual(request, documento):
    """Enviar email de reactivación a un cliente inactivo específico"""
    cliente = get_object_or_404(Cliente, documento=documento)
    
    if cliente.estado != 'inactivo':
        messages.warning(request, 'Este cliente está activo. No necesita reactivación.')
        return redirect('clientes_ver', documento=documento)
    
    if EmailService.enviar_email_reactivacion(cliente):
        messages.success(request, f'✓ Email de reactivación enviado a {cliente.email}')
    else:
        messages.error(request, '✗ Error al enviar el email')
    
    return redirect('clientes_ver', documento=documento)
# ========== EMAILS PARA CLIENTES INACTIVOS ==========

@login_required
@user_passes_test(es_administrador)
def emails_clientes_inactivos(request):
    """Panel para ver y enviar correos masivos a clientes inactivos"""
    
    # Obtener clientes inactivos con email
    clientes_inactivos = Cliente.objects.filter(
        estado='inactivo'
    ).exclude(email__isnull=True).exclude(email='')
    
    total_clientes = clientes_inactivos.count()
    
    context = {
        'clientes_inactivos': clientes_inactivos,
        'total_clientes': total_clientes,
    }
    
    return render(request, 'emails/clientes_inactivos.html', context)


@login_required
@user_passes_test(es_administrador)
def enviar_emails_inactivos(request):
    """Enviar emails masivos a clientes inactivos"""
    if request.method == 'POST':
        try:
            # Obtener clientes inactivos con email
            clientes_inactivos = Cliente.objects.filter(
                estado='inactivo'
            ).exclude(email__isnull=True).exclude(email='')
            
            if not clientes_inactivos.exists():
                messages.warning(request, 'No hay clientes inactivos con email registrado.')
                return redirect('emails_clientes_inactivos')
            
            # Enviar correos usando el EmailService
            correos_enviados = 0
            correos_fallidos = 0
            
            for cliente in clientes_inactivos:
                if EmailService.enviar_email_reactivacion(cliente):
                    correos_enviados += 1
                else:
                    correos_fallidos += 1
            
            # Mensajes de resultado
            if correos_enviados > 0:
                messages.success(
                    request, 
                    f'✓ Se enviaron {correos_enviados} correos exitosamente.'
                )
            
            if correos_fallidos > 0:
                messages.warning(request, f'⚠ {correos_fallidos} correos fallaron.')
            
            if correos_enviados == 0:
                messages.error(request, '✗ No se pudo enviar ningún correo.')
            
            return redirect('emails_clientes_inactivos')
            
        except Exception as e:
            messages.error(request, f'✗ Error al enviar correos: {str(e)}')
            return redirect('emails_clientes_inactivos')
    
    return redirect('emails_clientes_inactivos')


# ========== EMAILS INDIVIDUALES DE RENOVACIÓN ==========

@login_required
def enviar_email_renovacion_individual(request, documento):
    """Enviar email de renovación a un cliente específico"""
    cliente = get_object_or_404(Cliente, documento=documento)
    
    if EmailService.enviar_email_renovacion(cliente):
        messages.success(request, f'✓ Email de renovación enviado a {cliente.email}')
    else:
        messages.error(request, '✗ Error al enviar el email')
    
    return redirect('clientes_ver', documento=documento)

@login_required
def cliente_asistencias(request, documento):
    """Ver asistencias de un cliente con filtro por rango de fechas"""
    cliente = get_object_or_404(Cliente, documento=documento)
    
    # Obtener fechas del filtro
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    # Query base
    asistencias = Asistencia.objects.filter(cliente=cliente).order_by('-fecha', '-hora')
    
    # Aplicar filtros si existen
    if fecha_desde:
        asistencias = asistencias.filter(fecha__gte=fecha_desde)
    
    if fecha_hasta:
        asistencias = asistencias.filter(fecha__lte=fecha_hasta)
    
    # Estadísticas
    total_asistencias = asistencias.count()
    
    # Calcular rango de días
    if fecha_desde and fecha_hasta:
        fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        dias_rango = (fecha_hasta_obj - fecha_desde_obj).days + 1
    else:
        dias_rango = None
    
    context = {
        'cliente': cliente,
        'asistencias': asistencias,
        'total_asistencias': total_asistencias,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'dias_rango': dias_rango,
    }
    
    return render(request, 'clientes/asistencias.html', context)