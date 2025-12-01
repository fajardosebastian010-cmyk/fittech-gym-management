from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.http import HttpResponse
from .models import Cliente, Pago, Asistencia, Membresia
from django.db.models import Sum, Count
from io import BytesIO

# Colores corporativos FITTECH
COLOR_HEADER = "667EEA"  # Púrpura principal
COLOR_SUBHEADER = "764BA2"  # Púrpura oscuro
COLOR_ALT_ROW = "F3F4F6"  # Gris claro
COLOR_TOTAL = "10B981"  # Verde
COLOR_WARNING = "F59E0B"  # Naranja
COLOR_DANGER = "EF4444"  # Rojo

class ReportesExcel:
    
    @staticmethod
    def aplicar_estilos_header(ws, row, columnas):
        """Aplicar estilos al encabezado"""
        font_header = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, valor in enumerate(columnas, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = valor
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = alignment
            cell.border = border
        
        ws.row_dimensions[row].height = 25
    
    @staticmethod
    def aplicar_estilos_fila(ws, row, valores, alternado=False):
        """Aplicar estilos a una fila de datos"""
        font = Font(name='Calibri', size=11)
        fill = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type='solid') if alternado else PatternFill()
        alignment = Alignment(horizontal='left', vertical='center')
        border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        
        for col, valor in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = valor
            cell.font = font
            cell.fill = fill
            cell.alignment = alignment
            cell.border = border
            
            # Alineación específica para números
            if isinstance(valor, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if '.' in str(valor):
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '#,##0'
    
    @staticmethod
    def agregar_titulo(ws, titulo, subtitulo=""):
        """Agregar título y subtítulo profesional"""
        # Título
        ws.merge_cells('A1:E1')
        cell_titulo = ws['A1']
        cell_titulo.value = titulo
        cell_titulo.font = Font(name='Calibri', size=16, bold=True, color='667EEA')
        cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Subtítulo
        ws.merge_cells('A2:E2')
        cell_sub = ws['A2']
        cell_sub.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}" + (f" | {subtitulo}" if subtitulo else "")
        cell_sub.font = Font(name='Calibri', size=10, italic=True, color='6B7280')
        cell_sub.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18
        
        return 3  # Retorna la fila siguiente para datos
    
    @staticmethod
    def ajustar_ancho_columnas(ws):
        """Ajustar ancho automático de columnas"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # ============ REPORTE DE CLIENTES ============
    @staticmethod
    def generar_reporte_clientes(filtro_estado=None):
        """Generar reporte de clientes"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Título
        fila_inicio = ReportesExcel.agregar_titulo(ws, "REPORTE DE CLIENTES FITTECH")
        fila = fila_inicio + 1
        
        # Encabezados
        columnas = ['Documento', 'Nombre', 'Apellido', 'Teléfono', 'Email', 'Membresía', 'Estado', 'Fecha Fin', 'Teléfono Emergencia']
        ReportesExcel.aplicar_estilos_header(ws, fila, columnas)
        fila += 1
        
        # Datos
        clientes = Cliente.objects.all()
        if filtro_estado:
            clientes = clientes.filter(estado=filtro_estado)
        
        for idx, cliente in enumerate(clientes):
            valores = [
                cliente.documento,
                cliente.nombres,
                cliente.apellidos,
                cliente.telefono,
                cliente.email,
                cliente.membresia_actual.nombre if cliente.membresia_actual else 'N/A',
                cliente.estado.upper(),
                cliente.fecha_fin_membresia.strftime('%d/%m/%Y') if cliente.fecha_fin_membresia else 'N/A',
                cliente.telefono_emergencia or 'N/A'
            ]
            ReportesExcel.aplicar_estilos_fila(ws, fila, valores, alternado=(idx % 2 == 0))
            fila += 1
        
        # Resumen
        fila += 1
        ws.merge_cells(f'A{fila}:D{fila}')
        cell = ws[f'A{fila}']
        cell.value = f"TOTAL DE CLIENTES: {clientes.count()}"
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        ReportesExcel.ajustar_ancho_columnas(ws)
        
        return wb
    
    # ============ REPORTE DE PAGOS ============
    @staticmethod
    def generar_reporte_pagos(filtro_estado=None):
        """Generar reporte de pagos"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos"
        
        # Título
        fila_inicio = ReportesExcel.agregar_titulo(ws, "REPORTE DE PAGOS FITTECH")
        fila = fila_inicio + 1
        
        # Encabezados
        columnas = ['ID', 'Cliente', 'Concepto', 'Monto', 'Método', 'Tipo', 'Estado', 'Fecha', 'Usuario']
        ReportesExcel.aplicar_estilos_header(ws, fila, columnas)
        fila += 1
        
        # Datos
        pagos = Pago.objects.select_related('cliente', 'usuario_registro').all()
        if filtro_estado:
            pagos = pagos.filter(estado=filtro_estado)
        
        total_ingresos = 0
        for idx, pago in enumerate(pagos):
            valores = [
                pago.id,
                f"{pago.cliente.nombres} {pago.cliente.apellidos}",
                pago.concepto,
                pago.monto,
                pago.get_metodo_pago_display(),
                pago.tipo_pago.upper(),
                pago.estado.upper(),
                pago.fecha_pago.strftime('%d/%m/%Y %H:%M'),
                pago.usuario_registro.nombre
            ]
            ReportesExcel.aplicar_estilos_fila(ws, fila, valores, alternado=(idx % 2 == 0))
            total_ingresos += pago.monto
            fila += 1
        
        # Resumen
        fila += 1
        ws.merge_cells(f'A{fila}:C{fila}')
        cell = ws[f'A{fila}']
        cell.value = f"TOTAL PAGOS: {pagos.count()}"
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells(f'D{fila}:E{fila}')
        cell = ws[f'D{fila}']
        cell.value = f"TOTAL INGRESOS: ${total_ingresos:,.2f}"
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        ReportesExcel.ajustar_ancho_columnas(ws)
        
        return wb
    
    # ============ REPORTE DE ASISTENCIAS ============
    @staticmethod
    def generar_reporte_asistencias(fecha_inicio=None, fecha_fin=None):
        """Generar reporte de asistencias"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencias"
        
        # Título
        subtitulo = ""
        if fecha_inicio and fecha_fin:
            subtitulo = f"del {fecha_inicio} al {fecha_fin}"
        
        fila_inicio = ReportesExcel.agregar_titulo(ws, "REPORTE DE ASISTENCIAS FITTECH", subtitulo)
        fila = fila_inicio + 1
        
        # Encabezados
        columnas = ['Cliente', 'Documento', 'Teléfono', 'Fecha', 'Hora', 'Usuario Registro']
        ReportesExcel.aplicar_estilos_header(ws, fila, columnas)
        fila += 1
        
        # Datos
        asistencias = Asistencia.objects.select_related('cliente', 'usuario_registro').all()
        
        if fecha_inicio:
            asistencias = asistencias.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            asistencias = asistencias.filter(fecha__lte=fecha_fin)
        
        for idx, asistencia in enumerate(asistencias.order_by('-fecha')):
            valores = [
                f"{asistencia.cliente.nombres} {asistencia.cliente.apellidos}",
                asistencia.cliente.documento,
                asistencia.cliente.telefono,
                asistencia.fecha.strftime('%d/%m/%Y'),
                asistencia.hora.strftime('%H:%M:%S') if asistencia.hora else 'N/A',
                asistencia.usuario_registro.nombre
            ]
            ReportesExcel.aplicar_estilos_fila(ws, fila, valores, alternado=(idx % 2 == 0))
            fila += 1
        
        # Resumen
        fila += 1
        ws.merge_cells(f'A{fila}:E{fila}')
        cell = ws[f'A{fila}']
        cell.value = f"TOTAL ASISTENCIAS: {asistencias.count()}"
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        ReportesExcel.ajustar_ancho_columnas(ws)
        
        return wb
    
    # ============ REPORTE CONSOLIDADO ============
    @staticmethod
    def generar_reporte_consolidado():
        """Generar reporte consolidado del gimnasio"""
        wb = Workbook()
        
        # HOJA 1: RESUMEN GENERAL
        ws = wb.active
        ws.title = "Resumen General"
        
        fila = ReportesExcel.agregar_titulo(ws, "REPORTE CONSOLIDADO FITTECH")
        fila += 1
        
        # Datos
        total_clientes = Cliente.objects.count()
        clientes_activos = Cliente.objects.filter(estado='activo').count()
        clientes_inactivos = Cliente.objects.filter(estado='inactivo').count()
        total_asistencias = Asistencia.objects.count()
        total_pagos = Pago.objects.filter(estado='validado').aggregate(Sum('monto'))['monto__sum'] or 0
        pagos_pendientes = Pago.objects.filter(estado='pendiente').count()
        
        datos_resumen = [
            ['CLIENTES', '', ''],
            ['Total Clientes', total_clientes, ''],
            ['Clientes Activos', clientes_activos, ''],
            ['Clientes Inactivos', clientes_inactivos, ''],
            ['', '', ''],
            ['ASISTENCIAS', '', ''],
            ['Total Asistencias', total_asistencias, ''],
            ['', '', ''],
            ['PAGOS', '', ''],
            ['Total Ingresos', f"${total_pagos:,.2f}", ''],
            ['Pagos Pendientes', pagos_pendientes, ''],
        ]
        
        for idx, fila_datos in enumerate(datos_resumen):
            if fila_datos[0] and not fila_datos[1]:
                # Encabezado de sección
                ws.merge_cells(f'A{fila}:C{fila}')
                cell = ws[f'A{fila}']
                cell.value = fila_datos[0]
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill = PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                # Dato
                valores = fila_datos
                ReportesExcel.aplicar_estilos_fila(ws, fila, valores, alternado=(idx % 2 == 0))
            
            fila += 1
        
        ReportesExcel.ajustar_ancho_columnas(ws)
        
        return wb


# Funciones para descargar los reportes
def descargar_reporte_clientes(request, filtro_estado=None):
    """Vista para descargar reporte de clientes"""
    wb = ReportesExcel.generar_reporte_clientes(filtro_estado)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Clientes_{datetime.now().strftime("%d_%m_%Y")}.xlsx"'
    
    wb.save(response)
    return response


def descargar_reporte_pagos(request, filtro_estado=None):
    """Vista para descargar reporte de pagos"""
    wb = ReportesExcel.generar_reporte_pagos(filtro_estado)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Pagos_{datetime.now().strftime("%d_%m_%Y")}.xlsx"'
    
    wb.save(response)
    return response


def descargar_reporte_asistencias(request):
    """Vista para descargar reporte de asistencias"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    wb = ReportesExcel.generar_reporte_asistencias(fecha_inicio, fecha_fin)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Asistencias_{datetime.now().strftime("%d_%m_%Y")}.xlsx"'
    
    wb.save(response)
    return response


def descargar_reporte_consolidado(request):
    """Vista para descargar reporte consolidado"""
    wb = ReportesExcel.generar_reporte_consolidado()
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Consolidado_{datetime.now().strftime("%d_%m_%Y")}.xlsx"'
    
    wb.save(response)
    return response
